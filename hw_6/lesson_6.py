from openpyxl import Workbook
import csv
import json

# Задача №1
# Декодировать в строку байтовое значение. Затем в байтовый вид
# в кодировке 'Latin' и затем результат снова декодировать в строку

code_1 = b'r\xc3\xa9sum\xc3\xa9'


def func_cod_de_cod(code):
    de_code_1 = code.decode()
    code_latin_1 = de_code_1.encode(encoding='Latin 1')
    de_code_2 = code_latin_1.decode(encoding='Latin 1')
    return print(
        f'Значение байтовой строки :{code}\n'
        f'Значенте байтовой строки после декодирования: '
        f'{de_code_1}\n'
        f'Байтовый вид {de_code_1} в кодировке "Latina 1" : '
        f'{code_latin_1}\n'
        f'Декодирования полученного результата: {de_code_2}'
    )


func_cod_de_cod(code_1)

# Задача №2
# Ввести с клавиатуры 4 строки и сохранить в разные переменные.
# создать файл и записать в него 2 строки и закрыть файл.
# Затем открыть файл на редактирование и дозаписать оставшиеся 2 строки.
# В итоге должно быть 4 строки,каждая начинается с новой строки.


str_1 = input('Введите 1-ю строку: ')
str_2 = input('Введите 2-ю строку: ')
str_3 = input('Введите 3-ю строку: ')
str_4 = input('Введите 4-ю строку: ')

with open('String.txt', 'w') as file:
    file.write(f'{str_1}\n'
               f'{str_2}\n')

with open('String.txt', 'a') as file:
    file.write(f'{str_3}\n'
               f'{str_4}')

with open('String.txt', 'r') as file:
    for i in file:
        print(i)

#
# Задача №3
# Создать словарь из 6-ти элементов. Ключ 6-ти значное число
# а значение кортеж из 2-х элементов имя и возраст.
# Записать данные в словарь на диск в json-файл.

date_dict = {
    100001: ('Ibragim', 25),
    100002: ('David', 23),
    100003: ('Kasimir', 25),
    100004: ('Sigizmund', 30),
    100005: ('Karl', 25),
    100006: ('Vanek', 25)
}

with open('date_dict.json', 'w') as file:
    json.dump(date_dict, file)

# Задача №4
# Прочитать сохранённый json-файл и записать данные на диск в csv-файл,
# первой строкой которого озаглавив каждый столбец и добавив
# новый столбец "телефон"


with open('date_dict.json', 'r') as file:
    json_data = json.load(file)
    print(json_data)

with open('date_dict.csv', 'w', newline='') as file:
    cvs_writer = csv.writer(file)
    header = ['id', 'name', 'age', 'phone']
    cvs_writer.writerow(header)

    for i in json_data:
        id_i = i
        name = json_data.get(i)[0]
        age = json_data.get(i)[1]
        phone = f'(+33){i}'

        cvs_writer.writerow(
            (id_i,
             name,
             age,
             phone)
        )

#
# №5
# Прочитать сохранненый csv-файл и сохранить данные в excel - файл,
# кроме возрастата.

wb = Workbook()
wb1 = wb.create_sheet("People", 0)
wb1.sheet_properties.tabColor = "A626C3"

with open('date_dict.csv', 'r') as file:
    line_reader = csv.reader(file)

    count = 1
    row_1 = 1

    for x in line_reader:
        column_1 = 2
        row_2 = 2
        column_2 = 1
        del x[2]

        for y in x:
            wb1.cell(row=column_1, column=row_1, value=y)
            column_1 += 1
        row_1 += 1

    for i in range(1, row_1 - 1):
        wb1.cell(row=column_2, column=row_2, value=f'person{count}')
        count += 1
        row_2 += 1

wb.save('date_excel.xlsx')
wb.close()
